from openpyxl import Workbook  
import openpyxl  
import json
import os
from PyQt5.QtGui import QBrush, QColor, QDesktopServices, QFont, QImage, QPainter, QPixmap#需要安装库 pip3 install PyQt5
from PyQt5.QtCore import  QObject,  QRect,  QUrl,  Qt, pyqtSignal 
from PyQt5.QtWidgets import QAbstractItemView,QHeaderView,QApplication, QTableWidgetItem, QFrame, QGridLayout, QGroupBox, QLabel,QFileDialog, QStackedWidget, QHBoxLayout, QVBoxLayout

from qfluentwidgets.components import Dialog  # 需要安装库 pip install "PyQt-Fluent-Widgets[full]" -i https://pypi.org/simple/
from qfluentwidgets import ProgressRing, SegmentedWidget, TableWidget,CheckBox, DoubleSpinBox, HyperlinkButton,InfoBar, InfoBarPosition, NavigationWidget, Slider, SpinBox, ComboBox, LineEdit, PrimaryPushButton, PushButton ,StateToolTip, SwitchButton, TextEdit, Theme,  setTheme ,isDarkTheme,qrouter,NavigationInterface,NavigationItemPosition, EditableComboBox
from qfluentwidgets import FluentIcon as FIF
from qframelesswindow import FramelessWindow, StandardTitleBar

 

class Widget_prompt_dict(QFrame): # 术语字典界面


    def __init__(self, text: str, parent=None,configurator=None,user_interface_prompter=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        self.configurator = configurator
        self.user_interface_prompter = user_interface_prompter
        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(3) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['Raw', 'Translations', 'Note']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序
        self.tableView.setBorderVisible(True) # 开启显示边框功能，从而可以修改表格角半径
        self.tableView.setBorderRadius(8) # 将表格组件的边角半径设置为x像素，从而实现圆角效果。

        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('New Line')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第三列添加"删除空白行"按钮
        button = PushButton('Delete Blank Line')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 2, button)
        button.clicked.connect(self.delete_blank_row)



        # -----创建第1_1个组，添加多个组件-----
        box1_1 = QGroupBox()
        box1_1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_1 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('Import Dictionary', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置Export Dictionary按钮
        self.pushButton2 = PushButton('Export Dictionary', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置Empty the dictionary按钮
        self.pushButton3 = PushButton('Empty the dictionary', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置Save Dictionary按钮
        self.pushButton4 = PushButton('Save Dictionary', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout1_1.addWidget(self.pushButton1)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton2)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton3)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton4)
        box1_1.setLayout(layout1_1)




        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“译时提示”标签
        label3 = QLabel( flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label3.setText("Adding a hint dictionary")

        #设置“译时提示”显示
        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label4.setText("(Original text triggers, automatic glossary construction)")


        #设置“译时提示”开
        self.checkBox2 = CheckBox('Enabling features')
        self.checkBox2.stateChanged.connect(self.checkBoxChanged2)

        layout3.addWidget(label3)
        layout3.addWidget(self.label4)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.checkBox2)
        box3.setLayout(layout3)


        # 把内容添加到容器中
        container.addWidget(box3)    
        container.addWidget(self.tableView)
        container.addWidget(box1_1)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(20)     
        container.setContentsMargins(50, 70, 50, 30)      


    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    # 将条目添加到表格的辅助函数
    def add_to_table(self, srt, dst, info):
            row = self.tableView.rowCount() - 1 #获取表格的倒数行数
            self.tableView.insertRow(row)    # 在表格中插入一行
            self.tableView.setItem(row, 0, QTableWidgetItem(srt))
            self.tableView.setItem(row, 1, QTableWidgetItem(dst))
            if info:
                self.tableView.setItem(row, 2, QTableWidgetItem(info))
            #设置新行的高度与前一行相同
            self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

    #导入字典按钮
    def Importing_dictionaries(self):
        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'All Files (*)')

        if Input_File:
            print(f'[INFO] Selected documents: {Input_File}')
            # 获取文件后缀
            file_suffix = Input_File.split('.')[-1].lower()
            
            # 根据文件后缀执行不同操作
            if file_suffix == 'json':

                # 执行JSON文件的操作
                with open(Input_File, 'r', encoding="utf-8") as f:
                    dictionary = json.load(f)

                # 检查数据是列表还是字典
                if isinstance(dictionary, list):
                    for item in dictionary:
                        if item.get("srt", "") and item.get("dst", ""): # 提示字典格式

                            # 格式例
                            # [
                            #   {
                            #     "srt": "xxxx",
                            #     "dst": "xxxx",
                            #     "info": "xxx",
                            #   }
                            # ]

                            srt = item.get("srt", "")
                            dst = item.get("dst", "")
                            info = item.get("info", "")
                            self.add_to_table(srt, dst,info)

                        else: # Paratranz的术语表

                            # 格式例
                            # [
                            #   {
                            #     "id": 359894,
                            #     "createdAt": "2024-04-06T18:43:56.075Z",
                            #     "updatedAt": "2024-04-06T18:43:56.075Z",
                            #     "updatedBy": null,
                            #     "pos": "noun",
                            #     "uid": 49900,
                            #     "term": "アイテム",
                            #     "translation": "道具",
                            #     "note": "",
                            #     "project": 9841,
                            #     "variants": []
                            #   }
                            # ]

                            key = item.get("term", "")
                            value = item.get("translation", "")
                            info = ""
                            self.add_to_table(key, value,info)

                elif isinstance(dictionary, dict):  # 普通键值对格式
                    for key, value in dictionary.items():
                        info = ""
                        self.add_to_table(key, value,info)

                # 输出日志
                self.user_interface_prompter.createSuccessInfoBar("Imported successfully")
                print(f'[INFO]  Imported dictionary file')


            elif file_suffix == 'xlsx':
                # 执行XLSX文件的操作
                wb = openpyxl.load_workbook(Input_File)
                sheet = wb.active
                for row in range(2, sheet.max_row + 1): # 第一行是标识头，第二行才开始读取
                    cell_value1 = sheet.cell(row=row, column=1).value # 第N行第一列的值
                    cell_value2 = sheet.cell(row=row, column=2).value # 第N行第二列的值
                    cell_value3 = sheet.cell(row=row, column=3).value # 第N行第三列的值
                    self.add_to_table(cell_value1, cell_value2,cell_value3)

                # 输出日志
                self.user_interface_prompter.createSuccessInfoBar("Imported successfully")
                print(f'[INFO]  Imported dictionary file')
                
            else:
                print(f'[INFO] Unsupported file types: .{file_suffix}')

        else:
            print('[INFO] No document selected')
            return        
        
    #Export Dictionary按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间字典中
        dictionary = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            info_item = self.tableView.item(row, 2)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                if info_item:
                    info = info_item.text()
                    dictionary.append({"srt":key,"dst":value,"info":info})
                else:
                    dictionary.append({"srt":key,"dst":value})


        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  Dictionary export folder selected: {Output_Folder}')
        else :
            print('[INFO]  No folder selected')
            return  # 直接返回，不执行后续操作

        # 将字典保存到文件中
        with open(os.path.join(Output_Folder, "用户提示字典.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        self.user_interface_prompter.createSuccessInfoBar("Export Successful")
        print(f'[INFO]  Exported Dictionary file')

    #Empty the dictionary按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('New Line')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第三列添加"删除空白行"按钮
        button = PushButton('Delete Blank Line')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 2, button)
        button.clicked.connect(self.delete_blank_row)

        self.user_interface_prompter.createSuccessInfoBar("Empty successfully")
        print(f'[INFO]  已Empty the dictionary')

    #Save Dictionary按钮
    def Save_dictionary(self):
        self.user_interface_prompter.read_write_config("write",self.configurator.resource_dir)
        self.user_interface_prompter.createSuccessInfoBar("Save Successful")
        print(f'[INFO]  已Save Dictionary')

    
    #消息提示函数
    def checkBoxChanged2(self, isChecked: bool):
        if isChecked :
            self.user_interface_prompter.createSuccessInfoBar("Terminology Dictionary enabled")
